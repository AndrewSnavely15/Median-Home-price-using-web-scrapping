
from datetime import date
import openpyxl

from selenium import webdriver


driver = webdriver.Chrome(r'C:\Users\Andrew Snavely\Downloads\chromedriver_win32\chromedriver.exe')
driver.get("https://www.zillow.com/san-antonio-tx/home-values/")
driver.maximize_window()
price = driver.find_element_by_xpath("//h2[contains(text(), '$')]").text
driver.close()
print(str(price))
today = date.today()
today = today.strftime("%m/%d/%y")
print(today)

book = openpyxl.load_workbook(r'C:\Users\Andrew Snavely\Desktop\SA Home Prices.xlsx')

sheet = book.active

if sheet['A1'].value is None:
    book = openpyxl.load_workbook(r'C:\Users\Andrew Snavely\Desktop\SA Home Prices.xlsx')
    sheet = book.active
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Median Home Price'

    book.save(r'C:\Users\Andrew Snavely\Desktop\SA Home Prices.xlsx')
else:
    book = openpyxl.load_workbook(r'C:\Users\Andrew Snavely\Desktop\SA Home Prices.xlsx')
    sheet = book.active
    sheet.append([str(today), str(price)])
    book.save(r'C:\Users\Andrew Snavely\Desktop\SA Home Prices.xlsx')